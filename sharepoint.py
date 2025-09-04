# sharepoint.py
# Microsoft Graph API integration for SharePoint file operations with dynamic year tab names

import json
import requests
import os
import shutil
from pathlib import Path
from O365 import Account
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from config import (
    APPLICATION_CLIENT_ID,
    CLIENT_SECRET_VALUE,
    DIRECTORY_TENANT_ID,
    SITE_HOSTNAME,
    SITE_PATH,
    DEV_ROOT_FOLDER_ID,
    PROD_ROOT_FOLDER_ID
)


class SharePoint:
    def __init__(self):
        self.main_endpoint = "https://graph.microsoft.com/v1.0"
        self.host_name = SITE_HOSTNAME
        self.tenant_id = DIRECTORY_TENANT_ID
        self.client_id = APPLICATION_CLIENT_ID
        self.client_secret = CLIENT_SECRET_VALUE
        self.expiration_datetime = None
        self._site_url = "%s/sites/%s:/sites/%s"  # main_endpoint, host_name, site_path
        self.site_id = ""
        self.access_token = ""
        self.drive_id = ""

        # Define payer mapping to rows
        self.payer_row_mapping = {
            'Aetna': 5,
            'Amerigroup': 6,
            'Centene': 7,
            'CHPWA': 8,
            'Cigna': 9,
            'DSHS': 10,
            'HNB Echo': 11,
            'Humana': 12,
            'Kaiser': 13,
            'Medicare': 14,
            'Optum': 15,
            'Premera': 16,
            'Providence': 17,
            'Regence': 18,
            'Small Payers': 19,
            'Tricare': 20,
            'UHC': 21,
            'WA ST L&I': 22,
            'Zelis': 23
        }

        self.authenticate_and_get_drive_id()

    def get_month_sheet_name(self, year, month_number):
        """Generate dynamic sheet name for a specific year and month"""
        return f"{year}-{month_number}"

    def get_ytd_sheet_name(self, year):
        """Generate YTD sheet name for a specific year"""
        return f"{year}-YTD"

    def authenticate_and_get_drive_id(self):
        """This method authenticates the O365 account and gets the drive id."""
        self.access_token = self.authenticate_account()
        self.site_id = self.get_site_id(self.access_token)
        drive_info = self.get_site_drive_info(self.site_id, self.access_token)
        self.drive_id = drive_info["id"]

    def authenticate_account(self):
        """The function authenticates an O365 account using Microsoft Graph and returns the access token."""
        credentials = (self.client_id, self.client_secret)
        try:
            account = Account(credentials, auth_flow_type="credentials", tenant_id=self.tenant_id)
            account.authenticate()
            self.expiration_datetime = account.connection.token_backend.token.expiration_datetime
            with open("o365_token.txt") as f:
                data = f.read()
                js = json.loads(data)
                access_token = js["access_token"]
                print("Authentication successful!")
                return access_token
        except Exception as ex:
            print(f"Authentication failed: {ex}")
            raise ex

    def get_site_id(self, access_token):
        """The function retrieves the site ID for a given host name using an access token."""
        try:
            result = requests.get(
                self._site_url % (self.main_endpoint, self.host_name, SITE_PATH),
                headers={"Authorization": "Bearer " + access_token},
            )
            site_info = result.json()
            site_id = site_info["id"]
            print(f"Site ID retrieved: {site_id}")
            return site_id
        except Exception as ex:
            print(f"Getting Site ID Failed: {ex}")
            raise ex

    def get_site_drive_info(self, site_id, access_token):
        """The function retrieves drive information for a specific site using the site ID and access token."""
        try:
            result = requests.get(
                f"{self.main_endpoint}/sites/{site_id}/drive", headers={"Authorization": "Bearer " + access_token}
            )
            drive_info = result.json()
            print(f"Drive ID retrieved: {drive_info['id']}")
            return drive_info
        except Exception as ex:
            print(f"Getting Drive Info Failed: {ex}")
            raise ex

    def get_files_in_documents(self):
        """Get files from the Documents library"""
        try:
            result = requests.get(
                f"{self.main_endpoint}/drives/{self.drive_id}/root/children",
                headers={"Authorization": "Bearer " + self.access_token}
            )
            files_data = result.json()
            files = files_data.get('value', [])

            print(f"\nFound {len(files)} items in Documents library:")
            print("-" * 80)

            for file_item in files:
                file_type = 'FOLDER' if 'folder' in file_item else 'FILE'
                size_mb = file_item.get('size', 0) / (1024*1024) if file_item.get('size') else 0

                print(f"[{file_type}] {file_item.get('name')}")
                print(f"    ID: {file_item.get('id')}")
                print(f"    Size: {size_mb:.2f} MB")
                print(f"    Modified: {file_item.get('lastModifiedDateTime')}")
                if file_item.get('webUrl'):
                    print(f"    URL: {file_item.get('webUrl')}")
                print()

            return files
        except Exception as ex:
            print(f"Getting files failed: {ex}")
            raise ex

    def download_pmt_master_files(self, environment):
        """Download all PMT_MASTER files from the specified environment"""
        if environment not in ['dev', 'prod']:
            raise ValueError("Environment must be 'dev' or 'prod'")

        root_folder_id = DEV_ROOT_FOLDER_ID if environment == 'dev' else PROD_ROOT_FOLDER_ID
        print(f"Downloading PMT_MASTER files for {environment.upper()} environment...")

        local_dir = Path(environment)
        local_dir.mkdir(exist_ok=True)
        print(f"Created/verified local directory: {local_dir}")

        try:
            all_835s_folder = self._get_subfolder(root_folder_id, "All 835s")
            if not all_835s_folder:
                print("Could not find 'All 835s' folder")
                return []

            payer_folders = self._get_folder_contents(all_835s_folder['id'])
            payer_folders = [item for item in payer_folders if 'folder' in item]
            print(f"Found {len(payer_folders)} payer folders")

            downloaded_files = []

            for payer_folder in payer_folders:
                payer_name = payer_folder['name']
                print(f"Processing payer: {payer_name}")

                payer_files = self._get_folder_contents(payer_folder['id'])

                pmt_master_files = [
                    f for f in payer_files
                    if 'file' in f and f['name'].endswith(f"_PMT MASTER_{payer_name}.xlsx")
                ]

                print(f"  Found {len(pmt_master_files)} PMT_MASTER files")

                for file_item in pmt_master_files:
                    file_path = self._download_file(file_item['id'], file_item['name'], local_dir)
                    if file_path:
                        downloaded_files.append(file_path)
                        print(f"  Downloaded: {file_item['name']}")

            print(f"Total files downloaded: {len(downloaded_files)}")
            return downloaded_files

        except Exception as ex:
            print(f"Error downloading PMT_MASTER files: {ex}")
            raise ex

    def _get_subfolder(self, parent_folder_id, subfolder_name):
        """Get a specific subfolder by name"""
        try:
            contents = self._get_folder_contents(parent_folder_id)
            for item in contents:
                if 'folder' in item and item['name'] == subfolder_name:
                    return item
            return None
        except Exception as ex:
            print(f"Error getting subfolder '{subfolder_name}': {ex}")
            return None

    def _get_folder_contents(self, folder_id):
        """Get contents of a folder by ID"""
        try:
            result = requests.get(
                f"{self.main_endpoint}/drives/{self.drive_id}/items/{folder_id}/children",
                headers={"Authorization": "Bearer " + self.access_token}
            )
            result.raise_for_status()
            return result.json().get('value', [])
        except Exception as ex:
            print(f"Error getting folder contents: {ex}")
            raise ex

    def _download_file(self, file_id, file_name, local_dir):
        """Download a file by ID to local directory"""
        try:
            result = requests.get(
                f"{self.main_endpoint}/drives/{self.drive_id}/items/{file_id}/content",
                headers={"Authorization": "Bearer " + self.access_token}
            )
            result.raise_for_status()

            file_path = local_dir / file_name
            with open(file_path, 'wb') as f:
                f.write(result.content)

            return str(file_path)
        except Exception as ex:
            print(f"Error downloading file '{file_name}': {ex}")
            return None

    def _normalize_payment_note(self, note):
        """Map payment notes to normalized categories"""
        if pd.isna(note):
            return None

        note = str(note).strip()

        # Exclude these categories entirely
        if note in ['Proliance Backup Timeout', 'Batch Missing in NextGen']:
            return None

        # Map to normalized categories
        if note in ['Balanced-Batch Closed', 'Balanced-Batch Not Closed', 'Balanced']:
            return 'Balanced'
        elif note in ['Not Balanced-PLAs', 'Not Balanced-Remit Exceptions', 'Not Balanced-Expected']:
            return 'Not Balanced-Expected'
        elif note in ['Reconciled-Post Option Grayed Out', 'Not Balanced-Review', 'Not Balanced-TA Review']:
            return 'Not Balanced-Review'
        elif note == 'Amkai':
            return 'Amkai'
        else:
            return note

    def generate_monthly_stats(self, environment):
        """Process downloaded PMT MASTER files and generate Excel statistics reports"""
        stats_dir = Path(f"{environment}_stats")
        stats_dir.mkdir(exist_ok=True)

        files_dir = Path(environment)
        if not files_dir.exists():
            print(f"No {environment} folder found. Please download files first.")
            return

        excel_files = list(files_dir.glob("*_PMT MASTER_*.xlsx"))
        if not excel_files:
            print(f"No PMT MASTER files found in {environment} folder")
            return

        # Group files by year-month
        monthly_data = defaultdict(list)
        for file_path in excel_files:
            filename = file_path.name
            if filename.startswith(('2024-', '2025-', '2026-')):
                year_month = filename[:7]  # "2025-08"
                monthly_data[year_month].append(file_path)

        # Process both monthly and YTD reports
        if monthly_data:
            self._generate_reports_from_template(monthly_data, stats_dir)

        print(f"Excel reports generated in '{environment}_stats' folder")

    def _extract_payer_name(self, filename):
        """Extract payer name from filename"""
        try:
            parts = filename.split("_PMT MASTER_")
            if len(parts) == 2:
                return parts[1].replace('.xlsx', '')
        except:
            pass
        return None

    def _generate_reports_from_template(self, monthly_data, stats_dir):
        """Generate both monthly and YTD Excel reports using template"""
        years = set()
        for year_month in monthly_data.keys():
            year = year_month.split('-')[0]
            years.add(year)

        for year in sorted(years):
            year_months = [ym for ym in monthly_data.keys() if ym.startswith(year)]
            if not year_months:
                continue

            print(f"Generating Excel reports for {year}...")

            # Check if file already exists
            excel_file_path = stats_dir / f"{year}_Stats.xlsx"
            template_path = Path("Stats_template.xlsx")

            if not excel_file_path.exists():
                # Copy template to create new file
                if not template_path.exists():
                    print(f"Error: Stats_template.xlsx not found in root directory")
                    return

                print(f"Creating new stats file from template...")
                shutil.copy2(template_path, excel_file_path)

                # Update the template with correct year for ALL tabs
                wb = load_workbook(excel_file_path, keep_vba=False)

                # Update all worksheet names and content
                original_sheet_names = list(wb.sheetnames)  # Make a copy to avoid modification during iteration

                for original_sheet_name in original_sheet_names:
                    ws = wb[original_sheet_name]

                    # Update worksheet tab name
                    new_sheet_name = original_sheet_name.replace('YYYY', year)
                    if new_sheet_name != original_sheet_name:
                        ws.title = new_sheet_name
                        print(f"Renamed tab: {original_sheet_name} -> {new_sheet_name}")

                    # Update cell C3 which contains the year reference
                    if ws['C3'].value and 'YYYY' in str(ws['C3'].value):
                        ws['C3'].value = str(ws['C3'].value).replace('YYYY', year)
                        print(f"Updated C3 in {new_sheet_name}: {ws['C3'].value}")

                wb.save(excel_file_path)
                wb.close()
                print(f"Template copied and updated: {excel_file_path}")
            else:
                print(f"Stats file already exists: {excel_file_path}")

            # Process each month's data
            for year_month in year_months:
                year, month_number = year_month.split('-')  # "2025", "08"
                sheet_name = self.get_month_sheet_name(year, month_number)

                print(f"Processing data for {year_month} -> {sheet_name}")

                # Collect data for this specific month
                month_payer_stats = defaultdict(lambda: defaultdict(int))
                month_amkai_counts = defaultdict(int)

                files = monthly_data[year_month]
                for file_path in files:
                    payer_name = self._extract_payer_name(file_path.name)
                    if not payer_name:
                        continue

                    try:
                        df = pd.read_excel(file_path)
                        if df.empty:
                            continue

                        # Apply normalization to NOTE column
                        df['Normalized'] = df['NOTE'].apply(self._normalize_payment_note)

                        # Remove rows where normalization returned None
                        df_filtered = df[df['Normalized'].notna()]

                        # Count by normalized categories (excluding Amkai)
                        normalization_counts = df_filtered[df_filtered['Normalized'] != 'Amkai']['Normalized'].value_counts()

                        # Add to month totals
                        for category, count in normalization_counts.items():
                            month_payer_stats[payer_name][category] += count

                        # Count Amkai separately
                        amkai_count = (df_filtered['Normalized'] == 'Amkai').sum()
                        month_amkai_counts[payer_name] += amkai_count

                    except Exception as e:
                        print(f"Error processing {file_path.name}: {e}")
                        continue

                # Populate the monthly worksheet
                self._populate_monthly_worksheet(excel_file_path, sheet_name, month_payer_stats, month_amkai_counts)

            # Now generate YTD data by combining all months
            ytd_payer_stats = defaultdict(lambda: defaultdict(int))
            ytd_amkai_counts = defaultdict(int)

            for year_month in year_months:
                files = monthly_data[year_month]
                for file_path in files:
                    payer_name = self._extract_payer_name(file_path.name)
                    if not payer_name:
                        continue

                    try:
                        df = pd.read_excel(file_path)
                        if df.empty:
                            continue

                        df['Normalized'] = df['NOTE'].apply(self._normalize_payment_note)
                        df_filtered = df[df['Normalized'].notna()]
                        normalization_counts = df_filtered[df_filtered['Normalized'] != 'Amkai']['Normalized'].value_counts()

                        for category, count in normalization_counts.items():
                            ytd_payer_stats[payer_name][category] += count

                        amkai_count = (df_filtered['Normalized'] == 'Amkai').sum()
                        ytd_amkai_counts[payer_name] += amkai_count

                    except Exception as e:
                        continue

            # Populate YTD worksheet
            ytd_sheet_name = self.get_ytd_sheet_name(year)
            self._populate_monthly_worksheet(excel_file_path, ytd_sheet_name, ytd_payer_stats, ytd_amkai_counts)

    def _populate_monthly_worksheet(self, excel_path, sheet_name, payer_stats, amkai_counts):
        """Populate a specific worksheet with data"""
        print(f"Populating worksheet '{sheet_name}' in {excel_path}")

        try:
            wb = load_workbook(excel_path, data_only=False, keep_vba=False)

            if sheet_name not in wb.sheetnames:
                print(f"Warning: Worksheet '{sheet_name}' not found")
                print(f"Available worksheets: {wb.sheetnames}")
                wb.close()
                return

            ws = wb[sheet_name]

            # Populate data for each payer using correct columns: D, F, H, K
            for payer_name, stats in payer_stats.items():
                if payer_name in self.payer_row_mapping:
                    row = self.payer_row_mapping[payer_name]

                    try:
                        # Column D = Balanced
                        balanced_count = stats.get('Balanced', 0)
                        if ws[f'D{row}'].data_type != 'f':
                            ws[f'D{row}'].value = balanced_count

                        # Column F = Not Balanced-Expected
                        not_balanced_expected_count = stats.get('Not Balanced-Expected', 0)
                        if ws[f'F{row}'].data_type != 'f':
                            ws[f'F{row}'].value = not_balanced_expected_count

                        # Column H = Not Balanced-Review
                        not_balanced_review_count = stats.get('Not Balanced-Review', 0)
                        if ws[f'H{row}'].data_type != 'f':
                            ws[f'H{row}'].value = not_balanced_review_count

                        # Column K = Amkai
                        amkai_count = amkai_counts.get(payer_name, 0)
                        if ws[f'K{row}'].data_type != 'f':
                            ws[f'K{row}'].value = amkai_count

                        print(f"  {payer_name}: D{row}={balanced_count}, F{row}={not_balanced_expected_count}, H{row}={not_balanced_review_count}, K{row}={amkai_count}")
                    except Exception as e:
                        print(f"  Warning: Could not update {payer_name} - {e}")

            # Handle any payers not in our data (set to zero)
            for payer_name, row in self.payer_row_mapping.items():
                if payer_name not in payer_stats:
                    try:
                        if ws[f'D{row}'].data_type != 'f':
                            ws[f'D{row}'].value = 0
                        if ws[f'F{row}'].data_type != 'f':
                            ws[f'F{row}'].value = 0
                        if ws[f'H{row}'].data_type != 'f':
                            ws[f'H{row}'].value = 0
                        if ws[f'K{row}'].data_type != 'f':
                            ws[f'K{row}'].value = amkai_counts.get(payer_name, 0)
                    except Exception as e:
                        print(f"  Warning: Could not zero out {payer_name} - {e}")

            # Set cursor to N1
            try:
                ws.sheet_view.selection[0].activeCell = "N1"
                ws.sheet_view.selection[0].sqref = "N1"
            except:
                pass

            wb.save(excel_path)
            wb.close()
            print(f"Worksheet '{sheet_name}' populated successfully")

        except Exception as e:
            print(f"Error populating worksheet '{sheet_name}': {e}")
            raise


def main(environment):
    """Main function to run SharePoint operations"""
    try:
        client = SharePoint()

        print("=== ROOT DOCUMENTS ===")
        files = client.get_files_in_documents()

        print(f"\n=== DOWNLOADING PMT_MASTER FILES FOR {environment.upper()} ===")
        downloaded_files = client.download_pmt_master_files(environment)

        print(f"\nDownloaded {len(downloaded_files)} files to '{environment}' folder")

        if downloaded_files:
            print(f"\n=== GENERATING MONTHLY AND YTD STATISTICS FROM TEMPLATE ===")
            client.generate_monthly_stats(environment)

        return files
    except Exception as e:
        print(f"Error: {e}")
        return []


if __name__ == "__main__":
    main("prod")
    # main("dev")